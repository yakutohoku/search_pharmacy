"""
薬局マップ（Streamlit + Folium）

【主な機能】
1) Excel先頭シートの P列(緯度)・Q列(経度)を使って薬局を地図表示（青ピン）
   - 緯度/経度が空の行は表示しない
   - ピンをクリックすると
     薬局名(C列)、法人名（E列全文ラベル変更）、住所(I列) を表示
     Googleマップリンク・求人リンク（管理薬剤師K列, 常勤L列, パートM列, 派遣N列, 契約O列）

2) 検索
   - 住所・駅名で検索（半径指定 / ネットワークが必要）
   - 地図をクリックして指定（半径指定 / 確実）
   - 薬局名で検索（部分一致 + 都道府県で絞り込み）
   - 法人で検索（E列から法人名部分だけ抽出 + スペース無視で完全一致）
   - 社長名で検索（E列から「氏名」部分だけ抽出 + スペース無視で完全一致）

3) 薬局名で検索のときのみ
   - 一覧の店舗名ボタンをクリックすると、その薬局のピンが目立つ（色・大きさ変更）
   - 2km 圏内が見やすい倍率にズーム（2km円も表示）
   - ※薬局名検索では data_editor の一覧は表示しない

4) その他の検索モードでは
   - 複数選択（☑）→Excel出力 の機能を残す

【修正ポイント（今回）】
- 都道府県は「H列（都道府県）」を優先して読み込み（無い場合のみ住所から推定）
  → 都道府県絞り込みが正しく効くように修正

"""

from __future__ import annotations

import math
import re
import unicodedata
from dataclasses import dataclass
from io import BytesIO
from typing import Optional, Tuple, List, Dict, Any

import numpy as np
import pandas as pd
import streamlit as st
import folium
from folium.plugins import MarkerCluster
from streamlit_folium import st_folium
from branca.element import MacroElement, Template

# 住所 -> 緯度経度（外部通信が必要）
try:
    from geopy.geocoders import Nominatim, ArcGIS
    from geopy.extra.rate_limiter import RateLimiter
except Exception:
    Nominatim = None
    ArcGIS = None
    RateLimiter = None


# =============================================================================
# 設定
# =============================================================================
JOB_BASE_URL = "https://yaku-job.com/preview/"
GOOGLE_MAPS_QUERY_URL = "https://www.google.com/maps/search/?api=1&query={lat},{lon}"

# 参照情報（未読み込み時にメイン画面へ表示する）
REFERENCE_FOLDER = r"\\file-tky\Section\薬剤師共有\★【支店・課】_東北\東北\薬局検索"
REFERENCE_FILE_LIST = [
    "東北 薬局リスト.xlsm",
    "北海道 薬局リスト.xlsm",
]

# 都道府県（絞り込み対象）
PREFECTURE_SELECT = ["すべて", "北海道", "青森", "秋田", "岩手", "宮城", "山形", "福島"]
PREF_LABEL_TO_CANONICAL = {
    "北海道": "北海道",
    "青森": "青森県",
    "秋田": "秋田県",
    "岩手": "岩手県",
    "宮城": "宮城県",
    "山形": "山形県",
    "福島": "福島県",
}

# Excel列（0-based index）
# C=2, E=4, H=7, I=8, K=10, L=11, M=12, N=13, O=14, P=15, Q=16
COL = {
    "pharmacy_name": 2,    # C: 薬局名
    "opener_name": 4,      # E: 厚生局開設者氏名（法人名+代表取締役+氏名）
    "pref": 7,             # H: 都道府県（←今回ここを使う）
    "address": 8,          # I: 住所
    "id_manager": 10,      # K: 管理薬剤師ID
    "id_fulltime": 11,     # L: 常勤ID
    "id_part": 12,         # M: パートID
    "id_temp": 13,         # N: 派遣ID
    "id_contract": 14,     # O: 契約社員ID
    "lat": 15,             # P: 緯度
    "lon": 16,             # Q: 経度
}

DEFAULT_RADIUS_KM = 2.0
DEFAULT_MAX_MAP_MARKERS = 5000


# =============================================================================
# データ構造
# =============================================================================
@dataclass
class SearchPoint:
    lat: float
    lon: float


# =============================================================================
# 基本ユーティリティ
# =============================================================================
def _safe_float(x) -> Optional[float]:
    """Excelセル値を float へ安全に変換。変換不能/空/NaN/inf は None。"""
    if x is None:
        return None
    if isinstance(x, (int, float)):
        if isinstance(x, float) and (math.isnan(x) or math.isinf(x)):
            return None
        return float(x)

    s = str(x).strip().replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _normalize_id(x) -> Optional[str]:
    """求人IDをURL用に正規化（Excelの 123.0 対策など）。"""
    if x is None:
        return None
    s = str(x).strip()
    if not s or s.lower() in {"nan", "none"}:
        return None
    if s.endswith(".0"):
        s2 = s[:-2]
        if s2.isdigit():
            return s2
    return s


def _escape_html(text: str) -> str:
    """ポップアップHTMLの簡易エスケープ。"""
    return (text or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


# =============================================================================
# 検索用ユーティリティ
# =============================================================================
def normalize_space_ignored(text: str) -> str:
    """スペース（半角/全角）を無視して比較するための正規化（完全一致用）。"""
    t = unicodedata.normalize("NFKC", text or "")
    t = t.replace("　", " ")
    t = re.sub(r"\s+", "", t)
    return t


def normalize_contains_key(text: str) -> str:
    """部分一致（contains）用の正規化キー（空白を除去）。"""
    t = unicodedata.normalize("NFKC", text or "")
    t = t.replace("　", " ")
    t = re.sub(r"\s+", "", t)
    return t


def extract_corporation_part_from_opener(opener_raw: str) -> str:
    """
    E列（厚生局開設者氏名）から「法人名部分だけ」を切り出す。
    - 「代表取締役」等の手前までを法人名として採用
    """
    t = unicodedata.normalize("NFKC", opener_raw or "").replace("　", " ").strip()
    m = re.search(r"(代表取締役社長|代表取締役|代表社員|理事長|院長|代表者)", t)
    if m:
        t = t[: m.start()].strip()
    return t


def corporation_search_key_from_opener(opener_raw: str) -> str:
    """法人検索用キー（スペース無視で完全一致）。"""
    corp_part = extract_corporation_part_from_opener(opener_raw)
    return normalize_space_ignored(corp_part)


def extract_ceo_name_from_opener(opener_raw: str) -> str:
    """E列（厚生局開設者氏名）から「氏名部分」を切り出す。"""
    t = unicodedata.normalize("NFKC", opener_raw or "").replace("　", " ").strip()
    m = re.search(r"(代表取締役社長|代表取締役|代表社員|理事長|院長|代表者)\s*(.*)$", t)
    if not m:
        return ""
    name_part = (m.group(2) or "").strip()

    for prefix in ["社長", "会長", "CEO", "ＣＥＯ"]:
        if name_part.startswith(prefix):
            name_part = name_part[len(prefix):].strip()
    return name_part


def ceo_search_key_from_opener(opener_raw: str) -> str:
    """社長名検索キー（苗字・名前の間の空白も無視、完全一致）。"""
    name_part = extract_ceo_name_from_opener(opener_raw)
    return normalize_space_ignored(name_part)


def extract_prefecture_from_address(address: str) -> str:
    """
    住所から都道府県を推定（北海道・東北のみ）。
    ※今回は H列（都道府県）があるため、あくまでフォールバック
    """
    t = unicodedata.normalize("NFKC", address or "")
    t = t.replace("　", " ").strip()
    # 先頭の郵便番号を除去
    t = re.sub(r"^\s*〒?\s*\d{3}\s*-\s*\d{4}\s*", "", t)
    t = re.sub(r"^\s*〒?\s*\d{7}\s*", "", t).strip()

    for pref in ["北海道", "青森県", "秋田県", "岩手県", "宮城県", "山形県", "福島県"]:
        if pref in t:
            return pref

    if "北海道" in t:
        return "北海道"

    for short, canonical in [("青森", "青森県"), ("秋田", "秋田県"), ("岩手", "岩手県"), ("宮城", "宮城県"), ("山形", "山形県"), ("福島", "福島県")]:
        if short in t:
            return canonical

    return ""


def normalize_prefecture_value(pref_raw: Any, address_fallback: str) -> str:
    """
    H列（都道府県）を正規化して保存する
    - "青森" → "青森県" など
    - 空の場合のみ住所から推定
    """
    t = unicodedata.normalize("NFKC", str(pref_raw or "")).replace("　", " ").strip()
    if t and t.lower() not in {"nan", "none"}:
        # 既に正式表記っぽい場合
        if t == "北海道":
            return "北海道"
        if t.endswith("県") and len(t) <= 4:
            return t
        # "青森県八戸市..." のような混在でも拾う
        for canonical in ["青森県", "秋田県", "岩手県", "宮城県", "山形県", "福島県"]:
            if canonical in t:
                return canonical
        if "北海道" in t:
            return "北海道"
        # 県名省略
        for short, canonical in [("青森", "青森県"), ("秋田", "秋田県"), ("岩手", "岩手県"), ("宮城", "宮城県"), ("山形", "山形県"), ("福島", "福島県")]:
            if t == short or t.startswith(short):
                return canonical
        # ここまでで決められない場合も、含まれていれば採用
        for short, canonical in [("青森", "青森県"), ("秋田", "秋田県"), ("岩手", "岩手県"), ("宮城", "宮城県"), ("山形", "山形県"), ("福島", "福島県")]:
            if short in t:
                return canonical

    # フォールバック
    return extract_prefecture_from_address(address_fallback)


def prefecture_display(pref: str) -> str:
    """表示用（県を省略、北海道はそのまま）。"""
    p = (pref or "").strip()
    if not p:
        return "不明"
    if p == "北海道":
        return "北海道"
    return p.replace("県", "")


# =============================================================================
# UI（日本語化＆地図を大きく見せるCSS）
# =============================================================================
def _apply_ui_css() -> None:
    st.markdown(
        """
        <style>
        .block-container { padding-top: 1.2rem; padding-bottom: 1.0rem; }
        header[data-testid="stHeader"] { height: 0.2rem; }

        div[data-testid="stAppViewContainer"] h1 {
            font-size: 30px !important;
            font-weight: 800 !important;
            margin: 0.2rem 0 0.2rem 0 !important;
            line-height: 1.2 !important;
            overflow: visible !important;
            display: inline-block !important;
            max-width: 100% !important;
            white-space: normal !important;
        }

        /* uploader: 英文を消して日本語表示へ（DOM変更に備えて広めに当てる） */
        div[data-testid="stFileUploaderDropzone"] [data-testid="stFileUploaderDropzoneInstructions"] { display:none !important; }
        div[data-testid="stFileUploaderDropzone"] > div:nth-child(1) { display:none !important; }
        div[data-testid="stFileUploaderDropzone"] p { display:none !important; }
        div[data-testid="stFileUploaderDropzone"] small { display:none !important; }

        div[data-testid="stFileUploaderDropzone"]::before{
            content:"ここにExcelファイルをドラッグ＆ドロップ、または「ファイルを選択」を押してください";
            display:block;
            padding:0.25rem 0 0.6rem 0;
            font-size:0.95rem;
            font-weight:600;
            color:rgba(49,51,63,0.9);
        }

        div[data-testid="stFileUploaderDropzone"] button{
            position:relative;
        }
        div[data-testid="stFileUploaderDropzone"] button *{
            font-size:0px !important;
        }
        div[data-testid="stFileUploaderDropzone"] button::after{
            content:"ファイルを選択";
            font-size:14px;
            position:absolute;
            inset:0;
            display:flex;
            align-items:center;
            justify-content:center;
            pointer-events:none;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


# =============================================================================
# Excel読み込み（先頭シート固定）
# =============================================================================
@st.cache_data(show_spinner=False)
def load_pharmacy_data(file_bytes: bytes) -> pd.DataFrame:
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    if not wb.sheetnames:
        raise ValueError("Excelにシートが存在しません。")

    sheet = wb[wb.sheetnames[0]]

    records: List[Dict[str, Any]] = []

    for excel_row_no, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        lat = _safe_float(row[COL["lat"]] if len(row) > COL["lat"] else None)
        lon = _safe_float(row[COL["lon"]] if len(row) > COL["lon"] else None)
        if lat is None or lon is None:
            continue

        pharmacy_name = row[COL["pharmacy_name"]] if len(row) > COL["pharmacy_name"] else None
        opener_name = row[COL["opener_name"]] if len(row) > COL["opener_name"] else None
        pref_raw = row[COL["pref"]] if len(row) > COL["pref"] else None
        address = row[COL["address"]] if len(row) > COL["address"] else None

        opener_str = str(opener_name).strip() if opener_name is not None else ""
        pharmacy_str = str(pharmacy_name).strip() if pharmacy_name is not None else ""
        address_str = str(address).strip() if address is not None else ""

        corp_name = extract_corporation_part_from_opener(opener_str)
        ceo_name = extract_ceo_name_from_opener(opener_str)

        pref = normalize_prefecture_value(pref_raw, address_str)

        records.append(
            {
                "UID": str(excel_row_no),
                "Excel行番号": int(excel_row_no),

                "薬局名": pharmacy_str,
                "薬局名検索キー": normalize_contains_key(pharmacy_str),

                "開設者氏名": opener_str,
                "法人名": corp_name,
                "社長名": ceo_name,

                "法人検索キー": corporation_search_key_from_opener(opener_str),
                "社長検索キー": ceo_search_key_from_opener(opener_str),

                "住所": address_str,
                "都道府県": pref,

                "緯度": float(lat),
                "経度": float(lon),

                "管理薬剤師ID": _normalize_id(row[COL["id_manager"]] if len(row) > COL["id_manager"] else None),
                "常勤ID": _normalize_id(row[COL["id_fulltime"]] if len(row) > COL["id_fulltime"] else None),
                "パートID": _normalize_id(row[COL["id_part"]] if len(row) > COL["id_part"] else None),
                "派遣ID": _normalize_id(row[COL["id_temp"]] if len(row) > COL["id_temp"] else None),
                "契約社員ID": _normalize_id(row[COL["id_contract"]] if len(row) > COL["id_contract"] else None),
            }
        )

    df = pd.DataFrame.from_records(records)
    if df.empty:
        raise ValueError("緯度(P)・経度(Q)が入っている行が見つかりませんでした。")
    return df


# =============================================================================
# 地図：読み込みオーバーレイ（ブラウザ側）
# =============================================================================
def add_map_loading_overlay(m: folium.Map, message: str = "検索中・・・地図を読み込み中です") -> None:
    msg = _escape_html(message)
    map_name = m.get_name()

    tpl = Template(
        f"""
        {{% macro html(this, kwargs) %}}
        <style>
          .map-loading-overlay {{
            position: absolute;
            inset: 0;
            display: flex;
            align-items: center;
            justify-content: center;
            background: rgba(255,255,255,0.75);
            z-index: 9999;
            font-size: 16px;
            font-weight: 600;
            color: #333;
            pointer-events: none;
          }}
        </style>
        {{% endmacro %}}

        {{% macro script(this, kwargs) %}}
        (function() {{
          var map = {map_name};
          if (!map) return;

          var container = map.getContainer();
          if (!container) return;

          container.style.position = "relative";
          var overlay = document.createElement("div");
          overlay.className = "map-loading-overlay";
          overlay.innerText = "{msg}";
          container.appendChild(overlay);

          function show() {{ overlay.style.display = "flex"; }}
          function hide() {{ overlay.style.display = "none"; }}

          show();

          var hooked = false;
          map.eachLayer(function(layer) {{
            if (layer && layer instanceof L.TileLayer) {{
              hooked = true;
              layer.on("loading", show);
              layer.on("load", hide);
            }}
          }});

          if (!hooked) {{
            window.setTimeout(hide, 1500);
          }}
        }})();
        {{% endmacro %}}
        """
    )

    macro = MacroElement()
    macro._template = tpl
    m.get_root().add_child(macro)


# =============================================================================
# 地図：ピンのポップアップ
# =============================================================================
def _make_popup_html(r: pd.Series) -> str:
    name = _escape_html(str(r.get("薬局名", "")))
    opener_full = _escape_html(str(r.get("開設者氏名", "")))
    addr = _escape_html(str(r.get("住所", "")))

    lat = float(r["緯度"])
    lon = float(r["経度"])
    gmap = GOOGLE_MAPS_QUERY_URL.format(lat=lat, lon=lon)

    def job_li(label: str, job_id: Optional[str]) -> str:
        if job_id:
            return f'<li><a href="{JOB_BASE_URL + job_id}" target="_blank" rel="noopener noreferrer">{label}</a></li>'
        return f"<li>{label}: IDなし</li>"

    jobs = (
        "<div style='margin-top:6px;'><div><b>求人リンク</b></div>"
        "<ul style='margin:4px 0 0 18px;padding:0;'>"
        f"{job_li('管理薬剤師求人', r.get('管理薬剤師ID'))}"
        f"{job_li('常勤求人', r.get('常勤ID'))}"
        f"{job_li('パート求人', r.get('パートID'))}"
        f"{job_li('派遣求人', r.get('派遣ID'))}"
        f"{job_li('契約社員', r.get('契約社員ID'))}"
        "</ul></div>"
    )

    return f"""
    <div style="font-size:13px;line-height:1.4;max-width:420px;">
      <div style="font-size:14px;"><b>{name}</b></div>
      <div>法人名: {opener_full}</div>
      <div>住所: {addr}</div>
      <div style="margin-top:6px;">
        <a href="{gmap}" target="_blank" rel="noopener noreferrer">Googleマップを開く</a>
      </div>
      {jobs}
    </div>
    """


@st.cache_data(show_spinner=False)
def prepare_points_for_map(df: pd.DataFrame) -> List[Dict[str, Any]]:
    pts: List[Dict[str, Any]] = []
    for _, r in df.iterrows():
        pts.append(
            {
                "uid": str(r.get("UID", "")),
                "lat": float(r["緯度"]),
                "lon": float(r["経度"]),
                "tooltip": str(r.get("薬局名", "")),
                "popup_html": _make_popup_html(r),
            }
        )
    return pts


def build_map(
    center: Tuple[float, float],
    zoom: int,
    points: List[Dict[str, Any]],
    search_point: Optional[SearchPoint],
    radius_km: Optional[float],
    pending_point: Optional[SearchPoint],
    selected_uid: Optional[str],
    selected_circle_km: Optional[float],
) -> folium.Map:
    m = folium.Map(location=center, zoom_start=zoom, control_scale=True)

    if pending_point is not None:
        folium.Marker(
            location=(pending_point.lat, pending_point.lon),
            tooltip="選択中（未確定）",
            icon=folium.Icon(color="orange", icon="map-marker"),
        ).add_to(m)

    if search_point is not None:
        sp_html = f"""
        <div style="font-size:13px;line-height:1.4;">
          <div style="font-size:14px;"><b>検索地点</b></div>
          <div>緯度: {search_point.lat:.6f}</div>
          <div>経度: {search_point.lon:.6f}</div>
          <div style="margin-top:6px;">
            <a href="{GOOGLE_MAPS_QUERY_URL.format(lat=search_point.lat, lon=search_point.lon)}"
               target="_blank" rel="noopener noreferrer">Googleマップを開く</a>
          </div>
        </div>
        """
        folium.Marker(
            location=(search_point.lat, search_point.lon),
            tooltip="検索地点（赤ピン）",
            popup=folium.Popup(sp_html, max_width=360),
            icon=folium.Icon(color="red", icon="map-marker"),
        ).add_to(m)

        if radius_km and radius_km > 0:
            folium.Circle(
                location=(search_point.lat, search_point.lon),
                radius=radius_km * 1000,
                fill=False,
                weight=2,
            ).add_to(m)

    if selected_uid and selected_circle_km and selected_circle_km > 0:
        for p in points:
            if p.get("uid") == selected_uid:
                folium.Circle(
                    location=(p["lat"], p["lon"]),
                    radius=selected_circle_km * 1000,
                    fill=False,
                    weight=2,
                ).add_to(m)
                break

    cluster = MarkerCluster(name="薬局").add_to(m)
    for p in points:
        uid = p.get("uid")
        if selected_uid is not None and uid == selected_uid:
            folium.Marker(
                location=(p["lat"], p["lon"]),
                tooltip=p["tooltip"],
                popup=folium.Popup(p["popup_html"], max_width=520),
                icon=folium.Icon(color="red", icon="map-marker", prefix="fa"),
            ).add_to(cluster)
        else:
            folium.Marker(
                location=(p["lat"], p["lon"]),
                tooltip=p["tooltip"],
                popup=folium.Popup(p["popup_html"], max_width=480),
                icon=folium.Icon(color="blue", icon="info-sign"),
            ).add_to(cluster)

    folium.LayerControl().add_to(m)
    return m


# =============================================================================
# 住所 -> 緯度経度（外部通信が必要）
# =============================================================================
def _address_candidates(address: str) -> List[str]:
    a = unicodedata.normalize("NFKC", address).replace("　", " ").strip()
    a = re.sub(r"^\s*〒?\s*\d{3}\s*-\s*\d{4}\s*", "", a)
    a = re.sub(r"^\s*〒?\s*\d{7}\s*", "", a).strip()

    b = re.sub(r"(\d+)\s*丁目", r"\1-", a)
    b = re.sub(r"(\d+)\s*番", r"\1-", b)
    b = re.sub(r"(\d+)\s*号", r"\1", b)
    b = re.sub(r"-{2,}", "-", b).strip("- ").strip()

    out: List[str] = []
    for q in [a, b, b + " 日本"]:
        q = q.strip()
        if q and q not in out:
            out.append(q)
    return out


def _haversine_km_vec(lat1: float, lon1: float, lat2: np.ndarray, lon2: np.ndarray) -> np.ndarray:
    r = 6371.0088
    phi1 = np.deg2rad(lat1)
    phi2 = np.deg2rad(lat2)
    dphi = np.deg2rad(lat2 - lat1)
    dlambda = np.deg2rad(lon2 - lon1)
    a = (np.sin(dphi / 2) ** 2) + (np.cos(phi1) * phi2 * 0 + np.cos(phi1) * np.cos(phi2) * (np.sin(dlambda / 2) ** 2))
    return 2 * r * np.arcsin(np.sqrt(a))


def geocode_address(address: str, bias_df: Optional[pd.DataFrame] = None) -> Optional[SearchPoint]:
    if not address.strip():
        return None
    if Nominatim is None or ArcGIS is None or RateLimiter is None:
        return None

    nom = Nominatim(user_agent="tohoku-pharmacy-map-app/1.0", timeout=10)
    nom_geocode = RateLimiter(nom.geocode, min_delay_seconds=1, swallow_exceptions=True)

    arc = ArcGIS(timeout=10)
    arc_geocode = RateLimiter(arc.geocode, min_delay_seconds=1, swallow_exceptions=True)

    candidates: List[SearchPoint] = []

    for q in _address_candidates(address):
        loc = nom_geocode(q, country_codes="jp")
        if loc is not None:
            candidates.append(SearchPoint(lat=float(loc.latitude), lon=float(loc.longitude)))

        loc2 = arc_geocode(q)
        if loc2 is not None:
            candidates.append(SearchPoint(lat=float(loc2.latitude), lon=float(loc2.longitude)))

    if not candidates:
        return None

    if bias_df is None or bias_df.empty:
        return candidates[0]

    c_lat = float(bias_df["緯度"].mean())
    c_lon = float(bias_df["経度"].mean())
    lat2 = np.array([sp.lat for sp in candidates], dtype=float)
    lon2 = np.array([sp.lon for sp in candidates], dtype=float)
    d = _haversine_km_vec(c_lat, c_lon, lat2, lon2)
    best_i = int(np.argmin(d))
    return candidates[best_i]


# =============================================================================
# フィルタ＆一覧リンク
# =============================================================================
def filter_within_radius(df: pd.DataFrame, center: SearchPoint, radius_km: float) -> pd.DataFrame:
    lat2 = df["緯度"].to_numpy(dtype=float)
    lon2 = df["経度"].to_numpy(dtype=float)
    d = _haversine_km_vec(center.lat, center.lon, lat2, lon2)

    out = df.copy()
    out["距離(km)"] = d
    out = out[out["距離(km)"] <= radius_km].sort_values("距離(km)")
    return out


def filter_by_corporation(df: pd.DataFrame, corp_query_raw: str) -> pd.DataFrame:
    q = normalize_space_ignored(corp_query_raw)
    if not q:
        return df.iloc[0:0].copy()
    return df[df["法人検索キー"] == q].copy()


def filter_by_ceo_name(df: pd.DataFrame, ceo_query_raw: str) -> pd.DataFrame:
    q = normalize_space_ignored(ceo_query_raw)
    if not q:
        return df.iloc[0:0].copy()
    return df[df["社長検索キー"] == q].copy()


def filter_by_pharmacy_name(df: pd.DataFrame, name_query_raw: str, pref_label: str) -> pd.DataFrame:
    q = normalize_contains_key(name_query_raw)
    if not q:
        return df.iloc[0:0].copy()

    mask = df["薬局名検索キー"].str.contains(re.escape(q), na=False)

    if pref_label and pref_label != "すべて":
        canonical = PREF_LABEL_TO_CANONICAL.get(pref_label, "")
        if canonical:
            mask = mask & (df["都道府県"] == canonical)

    return df[mask].copy()


def add_link_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()

    out["Googleマップ"] = (
        "https://www.google.com/maps/search/?api=1&query="
        + out["緯度"].astype(str)
        + ","
        + out["経度"].astype(str)
    )

    out["管理薬剤師求人"] = np.where(out["管理薬剤師ID"].fillna("").astype(str) != "", JOB_BASE_URL + out["管理薬剤師ID"].astype(str), "")
    out["常勤求人"] = np.where(out["常勤ID"].fillna("").astype(str) != "", JOB_BASE_URL + out["常勤ID"].astype(str), "")
    out["パート求人"] = np.where(out["パートID"].fillna("").astype(str) != "", JOB_BASE_URL + out["パートID"].astype(str), "")
    out["派遣求人"] = np.where(out["派遣ID"].fillna("").astype(str) != "", JOB_BASE_URL + out["派遣ID"].astype(str), "")
    out["契約社員"] = np.where(out["契約社員ID"].fillna("").astype(str) != "", JOB_BASE_URL + out["契約社員ID"].astype(str), "")

    return out


# =============================================================================
# 選択データのExcel出力（複数選択用）
# =============================================================================
def build_selected_excel_bytes(selected_df: pd.DataFrame) -> bytes:
    out_cols = ["薬局名", "法人名", "住所"]
    export_df = selected_df[out_cols].copy()

    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="選択薬局")
    return bio.getvalue()


# =============================================================================
# 上部ステータス帯
# =============================================================================
def _status_bar_html(
    total: int,
    shown: int,
    mode_label: str,
    radius_km: Optional[float],
    sp: Optional[SearchPoint],
    corp_query: Optional[str],
    ceo_query: Optional[str],
    pharmacy_query: Optional[str],
    pref_label: Optional[str],
) -> str:
    if mode_label == "法人で検索":
        q = _escape_html(corp_query or "")
        detail = f"<b>法人検索</b>：{q}"
    elif mode_label == "社長名で検索":
        q = _escape_html(ceo_query or "")
        detail = f"<b>社長名検索</b>：{q}（スペース無視で完全一致）"
    elif mode_label == "薬局名で検索":
        q = _escape_html(pharmacy_query or "")
        pref = _escape_html(pref_label or "すべて")
        detail = f"<b>薬局名検索</b>：{q}&nbsp;&nbsp;&nbsp;<b>都道府県</b>：{pref}"
    else:
        sp_text = "（未指定）" if sp is None else f"緯度 {sp.lat:.5f} / 経度 {sp.lon:.5f}"
        r = "" if radius_km is None else f"<b>半径 {radius_km:.1f} km 以内</b>"
        detail = f"{r}&nbsp;&nbsp;&nbsp;<b>検索点</b>：{sp_text}"

    return f"""
    <div style="
        background:#E8F5E9;
        border:1px solid rgba(0,0,0,0.06);
        padding:10px 12px;
        border-radius:10px;
        font-size:14px;
        display:flex;
        gap:18px;
        flex-wrap:wrap;
        align-items:center;
    ">
      <div><b>読み込み完了</b>：{total:,}件（緯度・経度あり）</div>
      <div><b>検索モード</b>：{mode_label}</div>
      <div>{detail}</div>
      <div><b>表示件数</b>：{shown:,}件</div>
    </div>
    """


# =============================================================================
# メイン
# =============================================================================
def main() -> None:
    st.set_page_config(page_title="薬局マップ", layout="wide")
    _apply_ui_css()

    st.title("薬局検索マップ")
    st.caption("手順：1) Excelを読み込む → 2) 検索方法を選ぶ → 3) 地図と一覧で確認 → 4) ☑で選択しExcel出力")

    # session defaults
    st.session_state.setdefault("clicked_point", None)
    st.session_state.setdefault("pending_point", None)
    st.session_state.setdefault("corp_query_raw", "")
    st.session_state.setdefault("ceo_query_raw", "")
    st.session_state.setdefault("pharmacy_query_raw", "")
    st.session_state.setdefault("pref_label", "すべて")
    st.session_state.setdefault("selected_map_uid", None)

    st.session_state.setdefault("selected_uids", set())
    st.session_state.setdefault("list_editor_df", None)
    st.session_state.setdefault("list_editor_uids", [])

    st.session_state.setdefault("radius_km", float(DEFAULT_RADIUS_KM))
    st.session_state.setdefault("max_map_markers", int(DEFAULT_MAX_MAP_MARKERS))

    # -------------------------------------------------------------------------
    # 1. データ読み込み（サイドバー）
    # -------------------------------------------------------------------------
    st.sidebar.header("1. データを読み込む")
    st.sidebar.write("Excelファイルを選択してください。")

    uploaded = st.sidebar.file_uploader(
        "Excelファイル",
        type=["xlsm", "xlsx"],
        label_visibility="collapsed",
    )
    file_bytes: Optional[bytes] = uploaded.getvalue() if uploaded is not None else None

    # 未読み込み時：参照情報をメイン画面に表示
    if file_bytes is None:
        st.info("左の「1. データを読み込む」からExcelを読み込んでください。")

        st.markdown("### ★参照フォルダ")
        st.write("薬局データは以下のフォルダにあります。")
        st.code(REFERENCE_FOLDER, language="text")

        st.markdown("### ★リスト一覧")
        for f in REFERENCE_FILE_LIST:
            st.write(f"・{f}")

        st.stop()

    # -------------------------------------------------------------------------
    # Excel読み込み
    # -------------------------------------------------------------------------
    try:
        df = load_pharmacy_data(file_bytes)
    except Exception as e:
        st.error(f"読み込みに失敗しました: {e}")
        st.stop()

    # -------------------------------------------------------------------------
    # 2. 検索（サイドバー）
    # -------------------------------------------------------------------------
    st.sidebar.header("2. 検索")

    search_mode = st.sidebar.radio(
        "検索方法",
        [
            "住所・駅名で検索（半径指定）",
            "地図をクリックして指定（半径指定）",
            "薬局名で検索",
            "法人で検索",
            "社長名で検索",
        ],
        index=0,
    )

    # よく変えない設定は各ボタンの下へ
    if search_mode == "住所・駅名で検索（半径指定）":
        st.sidebar.caption("例：宮城県仙台市青葉区中央1-1-1 / 仙台駅")
        address = st.sidebar.text_input("住所・駅名", value="", placeholder="例：仙台駅")
        if st.sidebar.button("この住所で検索"):
            with st.spinner("住所・駅名から緯度・経度を取得しています..."):
                sp = geocode_address(address, bias_df=df)
            if sp is None:
                st.warning("位置情報を取得できませんでした（社内ネットワーク等でブロックの可能性）。")
            else:
                st.session_state.clicked_point = sp
                st.session_state.pending_point = None
                st.session_state.corp_query_raw = ""
                st.session_state.ceo_query_raw = ""
                st.session_state.pharmacy_query_raw = ""
                st.session_state.selected_map_uid = None

        st.session_state.radius_km = float(
            st.sidebar.number_input("半径（km）", 0.1, 200.0, float(st.session_state.radius_km), 0.5)
        )
        st.session_state.max_map_markers = int(
            st.sidebar.number_input("地図ピンの最大表示数", 500, 20000, int(st.session_state.max_map_markers), 500)
        )

    elif search_mode == "地図をクリックして指定（半径指定）":
        if st.sidebar.button("検索地点の確定を解除（全件表示）"):
            st.session_state.clicked_point = None
            st.session_state.pending_point = None
            st.session_state.corp_query_raw = ""
            st.session_state.ceo_query_raw = ""
            st.session_state.pharmacy_query_raw = ""
            st.session_state.selected_map_uid = None

        st.session_state.radius_km = float(
            st.sidebar.number_input("半径（km）", 0.1, 200.0, float(st.session_state.radius_km), 0.5)
        )
        st.session_state.max_map_markers = int(
            st.sidebar.number_input("地図ピンの最大表示数", 500, 20000, int(st.session_state.max_map_markers), 500)
        )

    elif search_mode == "薬局名で検索":
        st.sidebar.caption("薬局名は部分一致で検索できます。必要に応じて都道府県で絞り込みできます。")
        pharmacy_raw = st.sidebar.text_input("薬局名", value=st.session_state.pharmacy_query_raw, placeholder="例：さくら薬局")
        pref_label = st.sidebar.selectbox(
            "都道府県",
            PREFECTURE_SELECT,
            index=(PREFECTURE_SELECT.index(st.session_state.pref_label) if st.session_state.pref_label in PREFECTURE_SELECT else 0),
        )
        c1, c2 = st.sidebar.columns(2)
        with c1:
            if st.sidebar.button("この薬局名で検索"):
                st.session_state.pharmacy_query_raw = pharmacy_raw
                st.session_state.pref_label = pref_label
                st.session_state.clicked_point = None
                st.session_state.pending_point = None
                st.session_state.corp_query_raw = ""
                st.session_state.ceo_query_raw = ""
                st.session_state.selected_map_uid = None
        with c2:
            if st.sidebar.button("解除"):
                st.session_state.pharmacy_query_raw = ""
                st.session_state.pref_label = "すべて"
                st.session_state.selected_map_uid = None

        st.session_state.max_map_markers = int(
            st.sidebar.number_input("地図ピンの最大表示数", 500, 20000, int(st.session_state.max_map_markers), 500)
        )

    elif search_mode == "法人で検索":
        st.sidebar.caption("E列から「代表取締役」の前までを法人名として検索します（スペース無視・完全一致）")
        corp_raw = st.sidebar.text_input("法人名", value=st.session_state.corp_query_raw)
        c1, c2 = st.sidebar.columns(2)
        with c1:
            if st.sidebar.button("この法人で検索"):
                st.session_state.corp_query_raw = corp_raw
                st.session_state.clicked_point = None
                st.session_state.pending_point = None
                st.session_state.ceo_query_raw = ""
                st.session_state.pharmacy_query_raw = ""
                st.session_state.selected_map_uid = None
        with c2:
            if st.sidebar.button("解除"):
                st.session_state.corp_query_raw = ""
        st.session_state.max_map_markers = int(
            st.sidebar.number_input("地図ピンの最大表示数", 500, 20000, int(st.session_state.max_map_markers), 500)
        )

    else:
        st.sidebar.caption("E列の「代表取締役」の後ろを氏名として検索します（スペース無視・完全一致）")
        ceo_raw = st.sidebar.text_input("社長名（苗字+名前）", value=st.session_state.ceo_query_raw)
        c1, c2 = st.sidebar.columns(2)
        with c1:
            if st.sidebar.button("この社長名で検索"):
                st.session_state.ceo_query_raw = ceo_raw
                st.session_state.clicked_point = None
                st.session_state.pending_point = None
                st.session_state.corp_query_raw = ""
                st.session_state.pharmacy_query_raw = ""
                st.session_state.selected_map_uid = None
        with c2:
            if st.sidebar.button("解除"):
                st.session_state.ceo_query_raw = ""
        st.session_state.max_map_markers = int(
            st.sidebar.number_input("地図ピンの最大表示数", 500, 20000, int(st.session_state.max_map_markers), 500)
        )

    # -------------------------------------------------------------------------
    # 絞り込み
    # -------------------------------------------------------------------------
    corp_query_raw = st.session_state.corp_query_raw.strip()
    ceo_query_raw = st.session_state.ceo_query_raw.strip()
    pharmacy_query_raw = st.session_state.pharmacy_query_raw.strip()
    pref_label = st.session_state.pref_label
    selected_uid = st.session_state.selected_map_uid

    if search_mode == "法人で検索":
        show_df = filter_by_corporation(df, corp_query_raw) if corp_query_raw else df.iloc[0:0].copy()
        search_point = None
        mode_label = "法人で検索"
    elif search_mode == "社長名で検索":
        show_df = filter_by_ceo_name(df, ceo_query_raw) if ceo_query_raw else df.iloc[0:0].copy()
        search_point = None
        mode_label = "社長名で検索"
    elif search_mode == "薬局名で検索":
        show_df = filter_by_pharmacy_name(df, pharmacy_query_raw, pref_label) if pharmacy_query_raw else df.iloc[0:0].copy()
        search_point = None
        mode_label = "薬局名で検索"
    else:
        search_point = st.session_state.clicked_point
        mode_label = "半径検索"
        show_df = df
        if search_point is not None:
            show_df = filter_within_radius(df, search_point, float(st.session_state.radius_km))

    # -------------------------------------------------------------------------
    # ステータス帯
    # -------------------------------------------------------------------------
    st.markdown(
        _status_bar_html(
            total=len(df),
            shown=len(show_df),
            mode_label=mode_label,
            radius_km=(None if mode_label in {"法人で検索", "社長名で検索", "薬局名で検索"} else float(st.session_state.radius_km)),
            sp=search_point,
            corp_query=(corp_query_raw if mode_label == "法人で検索" else None),
            ceo_query=(ceo_query_raw if mode_label == "社長名で検索" else None),
            pharmacy_query=(pharmacy_query_raw if mode_label == "薬局名で検索" else None),
            pref_label=(pref_label if mode_label == "薬局名で検索" else None),
        ),
        unsafe_allow_html=True,
    )

    col_map, col_list = st.columns([5, 2], gap="large")

    # 地図中心・ズーム
    selected_circle_km: Optional[float] = None
    if mode_label == "薬局名で検索" and selected_uid:
        sel = show_df[show_df["UID"].astype(str) == str(selected_uid)]
        if not sel.empty:
            center = (float(sel.iloc[0]["緯度"]), float(sel.iloc[0]["経度"]))
            zoom = 14
            selected_circle_km = 2.0
        else:
            st.session_state.selected_map_uid = None
            selected_uid = None
            center = (float(show_df["緯度"].mean()), float(show_df["経度"].mean())) if not show_df.empty else (float(df["緯度"].mean()), float(df["経度"].mean()))
            zoom = 11
    else:
        center = (float(show_df["緯度"].mean()), float(show_df["経度"].mean())) if not show_df.empty else (float(df["緯度"].mean()), float(df["経度"].mean()))
        zoom = 11 if (mode_label in {"法人で検索", "社長名で検索"} or (mode_label == "半径検索" and search_point is not None) or mode_label == "薬局名で検索") else 8

    # 地図
    with col_map:
        st.subheader("地図")
        pending_point: Optional[SearchPoint] = st.session_state.pending_point

        map_df = show_df
        if len(map_df) > int(st.session_state.max_map_markers):
            st.warning(
                f"表示が重くなるため、地図のピンは先頭 {int(st.session_state.max_map_markers):,} 件に制限しています。"
                "（サイドバーの『地図ピンの最大表示数』で変更できます）"
            )
            map_df = map_df.head(int(st.session_state.max_map_markers))

        with st.spinner("検索中・・・地図を表示しています"):
            points = prepare_points_for_map(map_df)
            fmap = build_map(
                center=center,
                zoom=zoom,
                points=points,
                search_point=search_point,
                radius_km=(None if mode_label in {"法人で検索", "社長名で検索", "薬局名で検索"} else float(st.session_state.radius_km)),
                pending_point=pending_point,
                selected_uid=(str(selected_uid) if selected_uid else None),
                selected_circle_km=selected_circle_km,
            )
            add_map_loading_overlay(fmap, "検索中・・・地図を読み込み中です")
            map_data = st_folium(fmap, width=None, height=820, returned_objects=["last_clicked"], key="map")

        if search_mode == "地図をクリックして指定（半径指定）":
            clicked = (map_data or {}).get("last_clicked")
            if clicked and "lat" in clicked and "lng" in clicked:
                st.session_state.pending_point = SearchPoint(lat=float(clicked["lat"]), lon=float(clicked["lng"]))
            pending = st.session_state.pending_point
            if pending is not None:
                st.info(f"選択中（未確定）：緯度 {pending.lat:.6f} / 経度 {pending.lon:.6f}")
                if st.button("この地点を検索地点として確定する"):
                    st.session_state.clicked_point = pending
                    st.session_state.pending_point = None
                    st.session_state.corp_query_raw = ""
                    st.session_state.ceo_query_raw = ""
                    st.session_state.pharmacy_query_raw = ""
                    st.session_state.selected_map_uid = None

        if mode_label == "薬局名で検索" and pharmacy_query_raw and show_df.empty:
            st.warning("一致する薬局名が見つかりませんでした。")

    # 一覧
    with col_list:
        st.subheader("一覧")

        if mode_label == "薬局名で検索":
            st.caption("店舗名をクリックすると、地図上で選択した薬局が目立つ表示になり、2kmが見やすい倍率にズームします。")

            if not pharmacy_query_raw:
                st.info("サイドバーで薬局名を入力して検索してください。")
            elif show_df.empty:
                st.info("該当する薬局がありません。")
            else:
                tmp = show_df.copy()
                tmp["都道府県_表示"] = tmp["都道府県"].apply(prefecture_display)
                tmp = tmp.sort_values(["都道府県_表示", "薬局名", "住所"]).reset_index(drop=True)

                max_buttons = 200
                if len(tmp) > max_buttons:
                    st.info(f"一覧が多いため、上から {max_buttons} 件を表示しています（都道府県で絞ると探しやすいです）。")
                    tmp = tmp.head(max_buttons)

                for _, r in tmp.iterrows():
                    uid = str(r["UID"])
                    pref_disp = prefecture_display(r["都道府県"])
                    label = f"{r['薬局名']}"
                    is_selected = (st.session_state.selected_map_uid == uid)
                    btn_label = ("✅ " if is_selected else "") + label
                    if st.button(btn_label, key=f"pick_{uid}"):
                        st.session_state.selected_map_uid = uid
                        st.success(f"選択：{label}")

                if st.session_state.selected_map_uid:
                    st.markdown("---")
                    if st.button("選択を解除", key="clear_pick"):
                        st.session_state.selected_map_uid = None

            # 薬局名検索では data_editor を出さない
            return

        # ここから従来どおり（薬局名検索以外）
        view = add_link_columns(show_df).reset_index(drop=True)
        cols = ["☑", "薬局名", "法人名", "住所", "Googleマップ", "管理薬剤師求人", "常勤求人", "パート求人", "派遣求人", "契約社員"]

        if view.empty:
            st.info("表示対象がありません。")
            return

        display_uids = view["UID"].astype(str).tolist()
        base = view.drop(columns=["☑"], errors="ignore").copy()
        base.insert(0, "☑", [uid in st.session_state.selected_uids for uid in display_uids])
        base = base[cols].copy()

        if (st.session_state.list_editor_df is None) or (st.session_state.list_editor_uids != display_uids):
            st.session_state.list_editor_df = base
            st.session_state.list_editor_uids = display_uids
            st.session_state.pop("list_editor", None)

        edited = st.data_editor(
            st.session_state.list_editor_df,
            use_container_width=True,
            height=720,
            hide_index=True,
            column_config={
                "☑": st.column_config.CheckboxColumn("☑"),
                "Googleマップ": st.column_config.LinkColumn("Googleマップ", display_text="開く"),
                "管理薬剤師求人": st.column_config.LinkColumn("管理", display_text="開く"),
                "常勤求人": st.column_config.LinkColumn("常勤", display_text="開く"),
                "パート求人": st.column_config.LinkColumn("パート", display_text="開く"),
                "派遣求人": st.column_config.LinkColumn("派遣", display_text="開く"),
                "契約社員": st.column_config.LinkColumn("契約", display_text="開く"),
            },
            key="list_editor",
        )

        uids_for_zip = st.session_state.list_editor_uids
        checked_uids = {uid for uid, checked in zip(uids_for_zip, edited["☑"].tolist()) if checked}
        selected = set(st.session_state.selected_uids)
        selected -= set(uids_for_zip)
        selected |= checked_uids
        st.session_state.selected_uids = selected

        st.markdown("---")
        c1, c2 = st.columns([1, 1])
        with c1:
            st.write(f"☑ 選択中：{len(st.session_state.selected_uids):,}件")
        with c2:
            if st.button("☑選択をすべて解除"):
                st.session_state.selected_uids = set()
                st.session_state.list_editor_df = None
                st.session_state.list_editor_uids = []
                st.session_state.pop("list_editor", None)

        sel_df = df[df["UID"].astype(str).isin(st.session_state.selected_uids)].copy()
        if not sel_df.empty:
            xlsx_bytes = build_selected_excel_bytes(sel_df)
            st.download_button(
                label="選択した薬局をExcelでダウンロード",
                data=xlsx_bytes,
                file_name="選択薬局リスト.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            st.info("☑を付けると、ここからExcelをダウンロードできます。")


if __name__ == "__main__":
    main()
